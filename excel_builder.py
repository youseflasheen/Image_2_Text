"""
Excel Builder — generates professionally styled Excel workbooks
from extracted document data using openpyxl.
"""

import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from gemini_service import DocumentData

logger = logging.getLogger(__name__)

# ── Style constants ──────────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
_HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

_SUBHEADER_FONT = Font(name="Calibri", bold=True, color="1F4E79", size=11)
_SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

_COL_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_COL_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

_KEY_FONT = Font(name="Calibri", bold=True, size=11, color="2B3E50")
_KEY_FILL = PatternFill(start_color="EBF1F5", end_color="EBF1F5", fill_type="solid")

_VALUE_FONT = Font(name="Calibri", size=11)
_VALUE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

_ALT_VALUE_FILL = PatternFill(start_color="F7F9FB", end_color="F7F9FB", fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6D9"),
    right=Side(style="thin", color="B4C6D9"),
    top=Side(style="thin", color="B4C6D9"),
    bottom=Side(style="thin", color="B4C6D9"),
)

_META_FONT = Font(name="Calibri", italic=True, size=9, color="888888")


def _is_arabic(text: str) -> bool:
    """Check whether text contains Arabic characters."""
    for ch in text:
        if "\u0600" <= ch <= "\u06FF" or "\u0750" <= ch <= "\u077F":
            return True
    return False


def _cell_alignment(text: str) -> Alignment:
    """Return RTL alignment for Arabic text, LTR otherwise."""
    if _is_arabic(text):
        return Alignment(horizontal="right", vertical="center", readingOrder=2)
    return Alignment(horizontal="left", vertical="center")


# ── Public API ───────────────────────────────────────────────────────────────


def build_excel(data: DocumentData, output_path: str | Path) -> Path:
    """Create a styled Excel workbook from extracted document data.

    Parameters
    ----------
    data : DocumentData
        Structured extraction result from Gemini.
    output_path : str | Path
        Where to save the .xlsx file.

    Returns
    -------
    Path
        The saved file path.
    """
    output_path = Path(output_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # ── Title row ────────────────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = f"📄 {data.document_type.replace('_', ' ').title()}"
    title_cell.font = _HEADER_FONT
    title_cell.fill = _HEADER_FILL
    title_cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 36

    # ── Metadata row ─────────────────────────────────────────────────────
    ws.merge_cells("A2:B2")
    meta_cell = ws["A2"]
    meta_cell.value = (
        f"Language: {data.language}  •  "
        f"Extracted: {datetime.now().strftime('%Y-%m-%d %H:%M')}  •  "
        f"Fields: {len(data.fields)}"
    )
    meta_cell.font = _META_FONT
    meta_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── Column headers ───────────────────────────────────────────────────
    row = 4
    for col, label in [(1, "Field"), (2, "Value")]:
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _COL_HEADER_FONT
        cell.fill = _COL_HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

    # ── Data rows ────────────────────────────────────────────────────────
    for idx, field in enumerate(data.fields):
        r = row + 1 + idx
        is_alt = idx % 2 == 1

        # Key column
        key_cell = ws.cell(row=r, column=1, value=field.key)
        key_cell.font = _KEY_FONT
        key_cell.fill = _KEY_FILL if not is_alt else _ALT_VALUE_FILL
        key_cell.border = _THIN_BORDER
        key_cell.alignment = _cell_alignment(field.key)

        # Value column
        val_cell = ws.cell(row=r, column=2, value=field.value)
        val_cell.font = _VALUE_FONT
        val_cell.fill = _VALUE_FILL if not is_alt else _ALT_VALUE_FILL
        val_cell.border = _THIN_BORDER
        val_cell.alignment = _cell_alignment(field.value)

        ws.row_dimensions[r].height = 24

    # ── Auto-size columns ────────────────────────────────────────────────
    max_key_len = max((len(f.key) for f in data.fields), default=10)
    max_val_len = max((len(f.value) for f in data.fields), default=20)
    ws.column_dimensions["A"].width = min(max(max_key_len + 4, 20), 50)
    ws.column_dimensions["B"].width = min(max(max_val_len + 4, 30), 80)

    # ── Freeze header ────────────────────────────────────────────────────
    ws.freeze_panes = "A5"

    # ── Save ─────────────────────────────────────────────────────────────
    wb.save(str(output_path))
    logger.info("Excel file saved → %s", output_path)
    return output_path
